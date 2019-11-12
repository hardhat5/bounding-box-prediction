import csv
from openpyxl import load_workbook

wb = load_workbook("test.xlsx")
sheet = wb.active

i = 1
j = 1
final_data = []
with open('output.csv', 'r') as csvRead:
    reader = csv.reader(csvRead)
    for row in reader:
        data = []
        #data.append(sheet.cell(row=j,column=1).value)
        if(i%2!=0):
            data.append(sheet.cell(row=j,column=1).value)
            for i in range(4):
                data.append(row[i])
            print(data)
            final_data.append(data)
            j+=1
        i+=1


csvRead.close()

with open('final_output.csv', 'w') as csvWrite:
    writer = csv.writer(csvWrite)
    for row in final_data:
        writer.writerow(row)
